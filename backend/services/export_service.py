import os
from fpdf import FPDF
from openpyxl import Workbook
from datetime import datetime

EXPORT_FOLDER = "data/output"
os.makedirs(EXPORT_FOLDER, exist_ok=True)


def exporteer_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rittenoverzicht"

    ws.append(["Bestand", "Van", "Adres", "Afstand (km)", "Categorie"])
    for item in data["resultaten"]:
        ws.append([
            os.path.basename(data["bestand"]),
            item.get("van", "?"),
            item["adres"],
            item["afstand_km"],
            item["categorie"]
        ])

    ws.append([])
    ws.append(["Totaal kilometers", data["km_totaal"]])
    ws.append(["Woon-werk kilometers", data["km_woonwerk"]])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    bestandsnaam = f"ocr_export_{timestamp}.xlsx"
    pad = os.path.join(EXPORT_FOLDER, bestandsnaam)
    wb.save(pad)
    return bestandsnaam


def exporteer_pdf(data):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt="RitSync - Rittenrapport", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 10, txt=f"Bestand: {os.path.basename(data['bestand'])}", ln=True)
    pdf.ln(5)

    for item in data["resultaten"]:
        regel = f"Van: {item.get('van', '?')} -> {item['adres']} | {item['afstand_km']} km | {item['categorie']}"
        pdf.multi_cell(0, 8, regel)

    pdf.ln(5)
    pdf.set_font("Arial", style="B", size=10)
    pdf.cell(200, 10, txt=f"Totaal: {data['km_totaal']} km", ln=True)
    pdf.cell(200, 10, txt=f"Woon-werk: {data['km_woonwerk']} km", ln=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    bestandsnaam = f"ocr_export_{timestamp}.pdf"
    pad = os.path.join(EXPORT_FOLDER, bestandsnaam)
    pdf.output(pad)
    return bestandsnaam
