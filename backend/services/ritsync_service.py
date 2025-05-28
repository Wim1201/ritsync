import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
import os
import re
import openpyxl
from fpdf import FPDF
from datetime import datetime
from backend.services.google_service import get_distance_km

def bereken_kilometers(adressen):
    """
    Bereken de totale afstand tussen een reeks adressen met Google Maps.
    De adressen zijn geordend en vormen een keten.
    """
    totale_afstand = 0.0

    for i in range(len(adressen) - 1):
        vertrek = adressen[i]
        aankomst = adressen[i + 1]
        afstand = get_distance_km(vertrek, aankomst)
        totale_afstand += afstand

    return round(totale_afstand, 2)

ADDRESS_REGEX = re.compile(r"(\d{4}\s?[A-Z]{2})\s+([\w\s']+\d+)\b", re.IGNORECASE)
START_ADDRESS = "Dr. Kuyperstraat, Dongen"

def extract_address(text):
    matches = ADDRESS_REGEX.findall(text)
    if matches:
        return ["{} {}".format(p[1].strip(), p[0].strip()) for p in matches]
    return []

def process_txt_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        text = f.read()
    addresses = extract_address(text)
    results = []
    for addr in addresses:
        km = get_distance_km(START_ADDRESS, addr)
        results.append({
            "datum": extract_date_from_filename(filepath),
            "adres": addr,
            "afstand_km": km
        })
    return results

def extract_date_from_filename(filename):
    match = re.search(r'(\d{8})', filename)
    if match:
        try:
            return datetime.strptime(match.group(1), '%d%m%Y').strftime('%Y-%m-%d')
        except:
            pass
    return "Onbekend"

def export_to_excel(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Datum", "Adres", "Afstand (km)"])
    for row in data:
        ws.append([row["datum"], row["adres"], row["afstand_km"]])
    wb.save(output_path)

def export_to_pdf(data, output_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="RitSync Ritregistratie", ln=1, align='C')
    pdf.ln(10)
    for row in data:
        line = f"{row['datum']} | {row['adres']} | {row['afstand_km']} km"
        pdf.cell(200, 10, txt=line, ln=1)
    pdf.output(output_path)

def run_ritsync(txt_files, export_excel=True, export_pdf=False):
    all_data = []
    for file in txt_files:
        if os.path.exists(file):
            all_data.extend(process_txt_file(file))
        else:
            print(f"Bestand niet gevonden: {file}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if export_excel:
        excel_path = f"data/output/ritsync_{timestamp}.xlsx"
        export_to_excel(all_data, excel_path)
        print(f"Excel opgeslagen: {excel_path}")
    if export_pdf:
        pdf_path = f"data/output/ritsync_{timestamp}.pdf"
        export_to_pdf(all_data, pdf_path)
        print(f"PDF opgeslagen: {pdf_path}")
